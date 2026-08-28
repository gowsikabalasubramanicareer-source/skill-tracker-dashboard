from __future__ import annotations

import io
import os
import re
from datetime import datetime
from zoneinfo import ZoneInfo
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook
from streamlit_autorefresh import st_autorefresh

# ================================================================
# EDIT ONLY THIS SMALL CONFIG SECTION IF YOU WANT TO RENAME THINGS
# ================================================================
APP_TITLE = "Skill Tracker Dashboard"
APP_SUBTITLE = "Student Performance & Skill Development Monitoring"
ACADEMIC_YEAR = "2026–27"
SEMESTER = "Semester III"
REFRESH_SECONDS = 60

# Dashboard Monitoring Benchmark (configurable, not sourced from Google Sheets)
ATTENDANCE_AT_RISK = 65.0
ATTENDANCE_ON_TRACK = 80.0
SKILL_CRITICAL = 40.0
SKILL_ON_TRACK = 60.0
STRENGTH_BENCHMARK = 75.0
COMPOSITE_MIN_COMPONENTS = 3

# Friendly programme names shown in the dashboard
PROGRAMMES = {
    "B.Com – Semester III": "B.Com",
    "BCA – Semester III": "BCA",
}

# Local file detection keywords. Normally you do NOT need to edit these.
LOCAL_FILE_KEYWORDS = {
    "B.Com": ["bcom", "b.com"],
    "BCA": ["bca"],
}
# ================================================================

# EduLearn-inspired palette from the reference image
TEAL = "#56BDB1"
TEAL_DARK = "#329D92"
YELLOW = "#F6C84C"
PINK = "#EF5B78"
BLUE = "#42A5D5"
PURPLE = "#8C7AE6"
BG = "#F5F7FA"
TEXT = "#293241"
MUTED = "#7A869A"
BORDER = "#E6EAF0"
CARD_COLORS = [TEAL, YELLOW, PINK, BLUE, PURPLE]

st.set_page_config(page_title=APP_TITLE, page_icon="📊", layout="wide")
st_autorefresh(interval=REFRESH_SECONDS * 1000, key="live_refresh")

st.markdown(
    f"""
    <style>
      .stApp {{background:{BG}; color:{TEXT};}}
      .block-container {{padding-top:1.1rem; padding-bottom:2rem; max-width:1500px;}}
      section[data-testid="stSidebar"] {{background:#FFFFFF; border-right:1px solid {BORDER};}}
      .hero {{
        background:linear-gradient(120deg,{TEAL} 0%, #68C7BC 100%);
        border-radius:18px; padding:25px 28px; margin-bottom:18px; color:white;
        box-shadow:0 8px 22px rgba(50,157,146,.12);
      }}
      .hero h1 {{margin:0;font-size:2rem;line-height:1.15;color:white;}}
      .hero p {{margin:7px 0 0 0;color:#F6FFFD;font-size:1rem;}}
      .hero-meta {{margin-top:15px;font-size:.9rem;color:#EDFFFC;}}
      .metric-card {{
        background:#FFFFFF; border:1px solid {BORDER}; border-radius:15px;
        padding:15px 16px 14px 16px; min-height:118px;
        box-shadow:0 4px 12px rgba(41,50,65,.045);
      }}
      .metric-label {{color:{MUTED};font-size:.86rem;font-weight:600;margin-bottom:7px;}}
      .metric-value {{color:{TEXT};font-size:1.75rem;font-weight:750;line-height:1.1;}}
      .metric-note {{color:{MUTED};font-size:.76rem;margin-top:8px;}}
      .accent-line {{height:5px;border-radius:8px;margin:-15px -16px 12px -16px;}}
      .section-title {{font-size:1.25rem;font-weight:750;color:{TEXT};margin:6px 0 8px 0;}}
      .small-note {{color:{MUTED};font-size:.82rem;}}
      .status-pill {{display:inline-block;padding:5px 10px;border-radius:999px;font-size:.8rem;font-weight:700;}}
      div[data-testid="stDataFrame"] {{background:white;border-radius:14px;}}
      .stTabs [data-baseweb="tab-list"] {{gap:6px;}}
      .stTabs [data-baseweb="tab"] {{background:white;border-radius:9px;padding:7px 12px;}}
      .stTabs [aria-selected="true"] {{color:{TEAL_DARK} !important;}}

      .live-chip {{display:inline-flex;align-items:center;gap:7px;background:rgba(255,255,255,.18);border:1px solid rgba(255,255,255,.35);padding:6px 10px;border-radius:999px;font-size:.78rem;font-weight:750;color:white;}}
      .live-dot {{width:8px;height:8px;border-radius:50%;background:#D8FFF7;box-shadow:0 0 0 4px rgba(216,255,247,.18);}}
      .connection-bad {{display:inline-flex;align-items:center;gap:7px;background:#FFF1F4;border:1px solid #FFD5DE;color:#C63D5B;padding:6px 10px;border-radius:999px;font-size:.78rem;font-weight:750;}}
      .student-banner {{background:#FFFFFF;border:1px solid #E6EAF0;border-radius:16px;padding:18px 20px;margin:5px 0 14px 0;box-shadow:0 4px 14px rgba(41,50,65,.045);}}
      .insight-card {{background:#FFFFFF;border:1px solid #E6EAF0;border-radius:14px;padding:15px 16px;height:100%;}}
      .insight-title {{font-size:.82rem;font-weight:800;color:#7A869A;text-transform:uppercase;letter-spacing:.03em;margin-bottom:7px;}}
      .alert-risk {{background:#FFF1F4;border-left:5px solid #EF5B78;border-radius:12px;padding:12px 14px;margin:6px 0;}}
      .alert-attention {{background:#FFF9E7;border-left:5px solid #F6C84C;border-radius:12px;padding:12px 14px;margin:6px 0;}}
      .alert-ok {{background:#EAF9F6;border-left:5px solid #329D92;border-radius:12px;padding:12px 14px;margin:6px 0;}}
      .benchmark-box {{background:#FFFFFF;border:1px solid #E6EAF0;border-radius:14px;padding:14px 16px;margin:8px 0 12px 0;}}
    </style>
    """,
    unsafe_allow_html=True,
)


def text(v):
    return "" if v is None else str(v).strip()


def normal(v):
    return re.sub(r"\s+", " ", text(v)).lower()


def number(v):
    if v is None or text(v) == "":
        return np.nan
    try:
        return float(v)
    except Exception:
        try:
            return float(text(v).replace("%", "").replace(",", ""))
        except Exception:
            return np.nan


def percent(v):
    x = number(v)
    if pd.isna(x):
        return np.nan
    return x * 100 if -1.5 <= x <= 1.5 else x


def register_like(v):
    return len(re.sub(r"\D", "", text(v))) >= 8


class GoogleRowsBook:
    def __init__(self, sheet_map: dict[str, list[list]]):
        self.sheet_map = sheet_map
        self.sheetnames = list(sheet_map.keys())

    def rows_for(self, sheet_name):
        return self.sheet_map.get(sheet_name, [])


def sheet_rows(wb, sheet_name):
    if hasattr(wb, "rows_for"):
        return wb.rows_for(sheet_name)
    return [list(r) for r in wb[sheet_name].iter_rows(values_only=True)]


def match_sheet(wb, candidates):
    names = list(wb.sheetnames)
    for candidate in candidates:
        for name in names:
            if normal(candidate) == normal(name):
                return name
    for candidate in candidates:
        for name in names:
            if normal(candidate) in normal(name):
                return name
    return None


def find_header(rows, tokens, max_rows=12):
    best_row, best_score = None, -1
    for i, row in enumerate(rows[:max_rows]):
        line = " | ".join(normal(x) for x in row)
        score = sum(t.lower() in line for t in tokens)
        if score > best_score:
            best_row, best_score = i, score
    return best_row if best_score >= max(1, len(tokens) - 1) else None


def find_col(headers, include, exclude=()):
    for i, h in enumerate(headers):
        h = normal(h)
        if all(x.lower() in h for x in include) and not any(x.lower() in h for x in exclude):
            return i
    return None


def merge_frames(base, other):
    if other is None or other.empty:
        return base
    if base is None or base.empty:
        return other.copy()
    cols = [c for c in other.columns if c != "Student Name" or "Student Name" not in base.columns]
    return base.merge(other[cols], on="Register No", how="outer")


# -------------------- PARSERS --------------------
def parse_attendance(wb):
    s = match_sheet(wb, ["Attendance"])
    if not s:
        return pd.DataFrame()
    rows = sheet_rows(wb, s)
    h = find_header(rows, ["register", "name", "present"], 8)
    if h is None:
        return pd.DataFrame()
    hd = rows[h]
    ri = find_col(hd, ["register"])
    ni = find_col(hd, ["name"])
    roll_i = find_col(hd, ["roll"])
    di = find_col(hd, ["department"])
    ti = find_col(hd, ["total", "day"], ["present", "absent"])
    pi = find_col(hd, ["present"])
    ai = find_col(hd, ["absent"])
    p_i = next((i for i, x in enumerate(hd) if normal(x) in {"%", "attendance %", "attendance%"}), None)

    out = []
    for row in rows[h + 1 :]:
        if ri is None or ri >= len(row) or not register_like(row[ri]):
            continue
        total = number(row[ti]) if ti is not None and ti < len(row) else np.nan
        present = number(row[pi]) if pi is not None and pi < len(row) else np.nan
        att = percent(row[p_i]) if p_i is not None and p_i < len(row) else np.nan
        if pd.isna(att) and pd.notna(total) and total > 0 and pd.notna(present):
            att = present / total * 100
        out.append({
            "Register No": text(row[ri]),
            "Student Name": text(row[ni]) if ni is not None and ni < len(row) else "",
            "Roll No": text(row[roll_i]) if roll_i is not None and roll_i < len(row) else "",
            "Department": text(row[di]) if di is not None and di < len(row) else "",
            "Total Days": total,
            "Present Days": present,
            "Absent Days": number(row[ai]) if ai is not None and ai < len(row) else np.nan,
            "Attendance %": att,
        })
    return pd.DataFrame(out)


def parse_aptitude_assessment(wb):
    s = match_sheet(wb, ["Aptitude Assessment", "Aptitude  Assessment"])
    if not s:
        return pd.DataFrame()
    rows = sheet_rows(wb, s)
    h = find_header(rows, ["reg", "student", "accuracy"], 8)
    if h is None:
        return pd.DataFrame()
    hd = rows[h]
    ri = find_col(hd, ["reg"])
    ni = find_col(hd, ["student", "name"]) or find_col(hd, ["name"])
    si = find_col(hd, ["actual", "questions", "solved"])
    ci = find_col(hd, ["answered", "right"])
    wi = find_col(hd, ["answered", "wrong"])
    ai = find_col(hd, ["accuracy"])

    out = []
    for row in rows[h + 1 :]:
        if ri is None or ri >= len(row) or not register_like(row[ri]):
            continue
        out.append({
            "Register No": text(row[ri]),
            "Student Name": text(row[ni]) if ni is not None and ni < len(row) else "",
            "Aptitude Questions Solved": number(row[si]) if si is not None and si < len(row) else np.nan,
            "Aptitude Correct": number(row[ci]) if ci is not None and ci < len(row) else np.nan,
            "Aptitude Wrong": number(row[wi]) if wi is not None and wi < len(row) else np.nan,
            "Aptitude Accuracy %": percent(row[ai]) if ai is not None and ai < len(row) else np.nan,
        })
    return pd.DataFrame(out)


def parse_aptitude_tasks(wb):
    s = match_sheet(wb, ["Aptitude"])
    if not s:
        return pd.DataFrame()
    rows = sheet_rows(wb, s)
    h = find_header(rows, ["register", "name"], 8)
    if h is None:
        return pd.DataFrame()
    hd = rows[h]
    ri = find_col(hd, ["register"])
    ni = find_col(hd, ["name"])
    if ri is None:
        return pd.DataFrame()
    start = max(ri, ni if ni is not None else ri) + 1

    out = []
    for row in rows[h + 1 :]:
        if ri >= len(row) or not register_like(row[ri]):
            continue
        vals = [number(x) for x in row[start:]]
        vals = [x for x in vals if pd.notna(x) and 0 <= x <= 10]
        if vals:
            scale = 10 if max(vals) > 5 else 5
            score = np.mean(vals) / scale * 100
        else:
            score = np.nan
        out.append({"Register No": text(row[ri]), "Aptitude Task Score %": score})
    return pd.DataFrame(out)


def parse_communication(wb):
    s = match_sheet(wb, ["Communication"])
    if not s:
        return pd.DataFrame()
    rows = sheet_rows(wb, s)
    id_row = find_header(rows, ["register", "student name"], 8)
    if id_row is None:
        id_row = find_header(rows, ["register", "name"], 8)
    if id_row is None:
        return pd.DataFrame()

    metric_row, best = id_row, -1
    for i in range(id_row, min(id_row + 5, len(rows))):
        count = sum(normal(x) == "total" or normal(x).startswith("total/") for x in rows[i])
        if count > best:
            metric_row, best = i, count

    idh = rows[id_row]
    mh = rows[metric_row]
    ri = find_col(idh, ["register"])
    ni = find_col(idh, ["student", "name"])
    if ni is None:
        ni = find_col(idh, ["name"])
    totals = [i for i, x in enumerate(mh) if normal(x) == "total" or normal(x).startswith("total/")]
    cefrs = [i for i, x in enumerate(mh) if "cefr" in normal(x)]

    out = []
    for row in rows[max(id_row, metric_row) + 1 :]:
        if ri is None or ri >= len(row) or not register_like(row[ri]):
            continue
        vals = [number(row[i]) for i in totals if i < len(row)]
        vals = [x for x in vals if pd.notna(x) and x >= 0]
        cefr_vals = [text(row[i]) for i in cefrs if i < len(row) and text(row[i]) not in {"", "-"}]
        score = min(np.mean(vals) / 25 * 100, 100) if vals else np.nan
        out.append({
            "Register No": text(row[ri]),
            "Student Name": text(row[ni]) if ni is not None and ni < len(row) else "",
            "Communication Score %": score,
            "Latest CEFR": cefr_vals[-1] if cefr_vals else "",
        })
    return pd.DataFrame(out)


def parse_outof5_skill(wb, candidates, label):
    s = match_sheet(wb, candidates)
    if not s:
        return pd.DataFrame()
    rows = sheet_rows(wb, s)
    h = find_header(rows, ["student", "id"], 10)
    if h is None:
        h = find_header(rows, ["student", "name"], 10)
    if h is None:
        return pd.DataFrame()

    idh = rows[h]
    ri = find_col(idh, ["student", "id"])
    if ri is None:
        ri = find_col(idh, ["reg"])
    ni = find_col(idh, ["student", "name"])
    if ni is None:
        ni = find_col(idh, ["name"])
    if ri is None:
        return pd.DataFrame()

    # Find the rubric row that explicitly says out of 5 / marks out of 5.
    rubric_row, rubric_cols, best = None, [], 0
    for i in range(min(10, len(rows))):
        cols = [j for j, x in enumerate(rows[i]) if "out of 5" in normal(x)]
        if len(cols) > best:
            rubric_row, rubric_cols, best = i, cols, len(cols)
    if not rubric_cols:
        # Fallback: only examine columns after Student ID and ignore serial/link/text columns.
        rubric_cols = list(range(ri + 1, max(len(r) for r in rows[:10])))

    out = []
    start_row = max(h, rubric_row if rubric_row is not None else h) + 1
    for row in rows[start_row:]:
        if ri >= len(row) or not register_like(row[ri]):
            continue
        vals = []
        for j in rubric_cols:
            if j >= len(row):
                continue
            x = number(row[j])
            if pd.notna(x) and 0 <= x <= 5:
                vals.append(x)
        avg = np.mean(vals) if vals else np.nan
        out.append({
            "Register No": text(row[ri]),
            "Student Name": text(row[ni]) if ni is not None and ni < len(row) else "",
            f"{label} Score %": avg / 5 * 100 if pd.notna(avg) else np.nan,
            f"{label} Assessments": len(vals),
        })
    return pd.DataFrame(out)


def parse_dsa_sheet(wb, candidates, label):
    s = match_sheet(wb, candidates)
    if not s:
        return pd.DataFrame()
    rows = sheet_rows(wb, s)
    h = find_header(rows, ["register", "student", "questions completed"], 10)
    if h is None:
        return pd.DataFrame()
    hd = rows[h]
    ri = find_col(hd, ["register"])
    ni = find_col(hd, ["student", "name"])
    qi = find_col(hd, ["questions", "completed"])
    if ri is None:
        return pd.DataFrame()

    out = []
    for row in rows[h + 1 :]:
        if ri >= len(row) or not register_like(row[ri]):
            continue
        out.append({
            "Register No": text(row[ri]),
            "Student Name": text(row[ni]) if ni is not None and ni < len(row) else "",
            label: number(row[qi]) if qi is not None and qi < len(row) else np.nan,
        })
    return pd.DataFrame(out)



# -------------------- STUDENT-LEVEL DETAIL PARSERS --------------------
def pretty_header(v):
    if isinstance(v, datetime):
        return v.strftime("%d %b %Y")
    return re.sub(r"\\s+", " ", text(v)).strip()


def ffill_values(row, width=None):
    width = width or len(row)
    out, last = [], ""
    for j in range(width):
        v = pretty_header(row[j]) if j < len(row) else ""
        if v:
            last = v
        out.append(last)
    return out


def first_student_row(rows, reg_col, start=0):
    for i in range(start, len(rows)):
        if reg_col < len(rows[i]) and register_like(rows[i][reg_col]):
            return i
    return None


def parse_attendance_daily(wb):
    s = match_sheet(wb, ["Attendance"])
    if not s:
        return pd.DataFrame()
    rows = sheet_rows(wb, s)
    h = find_header(rows, ["register", "name", "present"], 8)
    if h is None:
        return pd.DataFrame()
    hd = rows[h]
    ri = find_col(hd, ["register"])
    if ri is None:
        return pd.DataFrame()
    meta_cols = {
        x for x in [
            find_col(hd, ["sl"]), find_col(hd, ["register"]), find_col(hd, ["roll"]),
            find_col(hd, ["name"]), find_col(hd, ["department"]),
            find_col(hd, ["total", "day"], ["present", "absent"]),
            find_col(hd, ["present"]), find_col(hd, ["absent"]),
            next((i for i, x in enumerate(hd) if normal(x) in {"%", "attendance %", "attendance%"}), None),
        ] if x is not None
    }
    out = []
    for row in rows[h + 1:]:
        if ri >= len(row) or not register_like(row[ri]):
            continue
        reg = text(row[ri])
        for j, hdr in enumerate(hd):
            if j in meta_cols or j >= len(row):
                continue
            label = pretty_header(hdr)
            status = text(row[j])
            if label and status:
                out.append({"Register No": reg, "Date / Session": label, "Status": status})
    return pd.DataFrame(out)


def parse_communication_detail(wb):
    s = match_sheet(wb, ["Communication"])
    if not s:
        return pd.DataFrame()
    rows = sheet_rows(wb, s)
    id_row = find_header(rows, ["register", "student name"], 8)
    if id_row is None:
        id_row = find_header(rows, ["register", "name"], 8)
    if id_row is None:
        return pd.DataFrame()

    metric_row, best = id_row, -1
    for i in range(id_row, min(id_row + 6, len(rows))):
        count = sum(normal(x) == "total" or "cefr" in normal(x) for x in rows[i])
        if count > best:
            metric_row, best = i, count

    idh = rows[id_row]
    mh = rows[metric_row]
    ri = find_col(idh, ["register"])
    ni = find_col(idh, ["student", "name"])
    if ni is None:
        ni = find_col(idh, ["name"])
    if ri is None:
        return pd.DataFrame()

    width = max(len(r) for r in rows[:metric_row + 1])
    contexts = [ffill_values(rows[k], width) for k in range(metric_row)]
    totals = [j for j, x in enumerate(mh) if normal(x) == "total" or normal(x).startswith("total/")]
    cefrs = [j for j, x in enumerate(mh) if "cefr" in normal(x)]
    first_metric = max(ri, ni if ni is not None else ri) + 1
    groups = []
    prev_end = first_metric - 1
    for ti in totals:
        ci = next((c for c in cefrs if c > ti), None)
        start = prev_end + 1
        end = ci if ci is not None else ti
        groups.append((start, ti, ci, end))
        prev_end = end

    data_start = first_student_row(rows, ri, metric_row + 1)
    if data_start is None:
        return pd.DataFrame()

    out = []
    for row in rows[data_start:]:
        if ri >= len(row) or not register_like(row[ri]):
            continue
        reg = text(row[ri])
        name = text(row[ni]) if ni is not None and ni < len(row) else ""
        for start, ti, ci, end in groups:
            ctx = []
            for k in range(metric_row):
                v = contexts[k][ti] if ti < len(contexts[k]) else ""
                if v and v not in ctx and "register" not in normal(v) and "student" not in normal(v):
                    ctx.append(v)
            activity = max(ctx, key=len) if ctx else f"Activity {len(out)+1}"
            date = next((v for v in ctx if re.search(r"\\d{1,2}[/\\-]\\d{1,2}|\\d{4}", v)), "")
            criteria = []
            for j in range(start, ti):
                if j >= len(mh) or j >= len(row):
                    continue
                m = pretty_header(mh[j])
                v = text(row[j])
                if m and v != "":
                    criteria.append(f"{m}: {v}")
            total_v = number(row[ti]) if ti < len(row) else np.nan
            cefr_v = text(row[ci]) if ci is not None and ci < len(row) else ""
            out.append({
                "Register No": reg,
                "Student Name": name,
                "Activity": activity,
                "Date / Context": date,
                "Criteria Breakdown": " | ".join(criteria),
                "Total": total_v,
                "CEFR": cefr_v,
            })
    return pd.DataFrame(out)


def parse_aptitude_assessment_detail(wb):
    s = match_sheet(wb, ["Aptitude Assessment", "Aptitude  Assessment"])
    if not s:
        return pd.DataFrame(), pd.DataFrame()
    rows = sheet_rows(wb, s)
    h = find_header(rows, ["reg", "student", "accuracy"], 8)
    if h is None:
        return pd.DataFrame(), pd.DataFrame()
    hd = rows[h]
    ri = find_col(hd, ["reg"])
    ni = find_col(hd, ["student", "name"])
    email_i = find_col(hd, ["email"])
    planned_i = find_col(hd, ["total", "questions", "semester"])
    solved_i = find_col(hd, ["actual", "questions", "solved"])
    correct_i = find_col(hd, ["answered", "right"])
    wrong_i = find_col(hd, ["answered", "wrong"])
    accuracy_i = find_col(hd, ["accuracy"])
    assessment_avg_i = find_col(hd, ["assessment", "average"])
    if ri is None:
        return pd.DataFrame(), pd.DataFrame()

    data_start = first_student_row(rows, ri, h + 1)
    if data_start is None:
        return pd.DataFrame(), pd.DataFrame()

    summaries = []
    for row in rows[data_start:]:
        if ri >= len(row) or not register_like(row[ri]):
            continue
        summaries.append({
            "Register No": text(row[ri]),
            "Student Name": text(row[ni]) if ni is not None and ni < len(row) else "",
            "Email": text(row[email_i]) if email_i is not None and email_i < len(row) else "",
            "Questions Planned": number(row[planned_i]) if planned_i is not None and planned_i < len(row) else np.nan,
            "Questions Solved": number(row[solved_i]) if solved_i is not None and solved_i < len(row) else np.nan,
            "Correct Answers": number(row[correct_i]) if correct_i is not None and correct_i < len(row) else np.nan,
            "Wrong Answers": number(row[wrong_i]) if wrong_i is not None and wrong_i < len(row) else np.nan,
            "Overall Accuracy %": percent(row[accuracy_i]) if accuracy_i is not None and accuracy_i < len(row) else np.nan,
            "Overall Assessment Average %": percent(row[assessment_avg_i]) if assessment_avg_i is not None and assessment_avg_i < len(row) else np.nan,
        })

    # Continuous-assessment blocks start after the summary columns.
    group_row = rows[h]
    metric_row = h + 2 if h + 2 < data_start else data_start - 1
    date_row = h + 1 if h + 1 < data_start else None
    starts = [j for j in range(10, len(group_row)) if text(group_row[j])]
    ca = []
    for row in rows[data_start:]:
        if ri >= len(row) or not register_like(row[ri]):
            continue
        reg = text(row[ri])
        name = text(row[ni]) if ni is not None and ni < len(row) else ""
        for idx, start in enumerate(starts):
            end = starts[idx + 1] if idx + 1 < len(starts) else min(len(row), len(rows[metric_row]))
            assessment = pretty_header(group_row[start])
            date_val = pretty_header(rows[date_row][start]) if date_row is not None and start < len(rows[date_row]) else ""
            vals = {}
            for j in range(start, end):
                if j >= len(rows[metric_row]) or j >= len(row):
                    continue
                metric = pretty_header(rows[metric_row][j])
                if metric:
                    vals[metric] = row[j]
            solved = next((number(v) for k, v in vals.items() if "actual" in normal(k) and "solved" in normal(k)), np.nan)
            correct = next((number(v) for k, v in vals.items() if "answered right" in normal(k)), np.nan)
            totalq = next((number(v) for k, v in vals.items() if "total" in normal(k) and "question" in normal(k)), np.nan)
            mark_raw = next((v for k, v in vals.items() if "marks" in normal(k) and "%" in text(k)), None)
            assessed = pd.notna(solved) or pd.notna(correct)
            ca.append({
                "Register No": reg,
                "Student Name": name,
                "Assessment": assessment,
                "Date": date_val if normal(date_val) != "date" else "",
                "Total Questions": totalq,
                "Questions Solved": solved,
                "Correct Answers": correct,
                "Marks %": percent(mark_raw) if assessed and mark_raw is not None else np.nan,
                "Assessment Status": "Assessed" if assessed else "Not Assessed",
            })
    return pd.DataFrame(summaries), pd.DataFrame(ca)


def parse_aptitude_task_detail(wb):
    s = match_sheet(wb, ["Aptitude"])
    if not s:
        return pd.DataFrame()
    rows = sheet_rows(wb, s)
    h = find_header(rows, ["register", "name"], 8)
    if h is None:
        return pd.DataFrame()
    hd = rows[h]
    ri = find_col(hd, ["register"])
    ni = find_col(hd, ["name"])
    if ri is None:
        return pd.DataFrame()
    data_start = first_student_row(rows, ri, h + 1)
    if data_start is None:
        return pd.DataFrame()
    start_col = max(ri, ni if ni is not None else ri) + 1
    out = []
    for row in rows[data_start:]:
        if ri >= len(row) or not register_like(row[ri]):
            continue
        reg = text(row[ri])
        name = text(row[ni]) if ni is not None and ni < len(row) else ""
        for j in range(start_col, len(row)):
            result = text(row[j])
            if result == "":
                continue
            date_val = pretty_header(hd[j]) if j < len(hd) else ""
            meta = [pretty_header(rows[k][j]) for k in range(h + 1, data_start) if j < len(rows[k]) and text(rows[k][j])]
            task = next((v for v in meta if re.search(r"quiz|task|test|assignment|assessment", v, re.I)), "")
            category = next((v for v in meta if normal(v) in {"home task", "assessment", "class task", "home", "class"}), "")
            topic_candidates = [v for v in meta if v not in {task, category, "NA"} and not re.fullmatch(r"[0-9.]+", v)]
            topic = next((v for v in topic_candidates if len(v) > 2), "")
            nums = [number(v) for v in meta]
            nums = [v for v in nums if pd.notna(v) and 0 < v <= 10]
            max_score = nums[-1] if nums else 5.0
            score = number(row[j])
            status = "Absent" if normal(result) in {"ab", "abs", "absent"} else ("Assessed" if pd.notna(score) else result)
            out.append({
                "Register No": reg,
                "Student Name": name,
                "Date": date_val,
                "Task / Assessment": task,
                "Topic": topic,
                "Category": category,
                "Score": score,
                "Max Score": max_score,
                "Status": status,
            })
    return pd.DataFrame(out)


def parse_bcom_skill_detail(wb, candidates, label):
    s = match_sheet(wb, candidates)
    if not s:
        return pd.DataFrame(), pd.DataFrame()
    rows = sheet_rows(wb, s)
    h = find_header(rows, ["student", "id"], 10)
    if h is None:
        h = find_header(rows, ["student", "name"], 10)
    if h is None:
        return pd.DataFrame(), pd.DataFrame()
    idh = rows[h]
    ri = find_col(idh, ["student", "id"])
    if ri is None:
        ri = find_col(idh, ["reg"])
    ni = find_col(idh, ["student", "name"])
    if ni is None:
        ni = find_col(idh, ["name"])
    if ri is None:
        return pd.DataFrame(), pd.DataFrame()

    rubric_row, score_cols, best = None, [], 0
    for i in range(min(10, len(rows))):
        cols = [j for j, x in enumerate(rows[i]) if "out of 5" in normal(x)]
        if len(cols) > best:
            rubric_row, score_cols, best = i, cols, len(cols)
    if rubric_row is None:
        return pd.DataFrame(), pd.DataFrame()
    data_start = first_student_row(rows, ri, rubric_row + 1)
    if data_start is None:
        return pd.DataFrame(), pd.DataFrame()
    width = max(len(r) for r in rows[:rubric_row + 1])
    ffilled = [ffill_values(rows[k], width) for k in range(rubric_row)]

    # Report-link columns can sit on a different header row from the ID row.
    report_cols = set()
    for k in range(rubric_row + 1):
        for j, v in enumerate(rows[k]):
            if "report link" in normal(v):
                report_cols.add(j)

    sessions, reports = [], []
    for row in rows[data_start:]:
        if ri >= len(row) or not register_like(row[ri]):
            continue
        reg = text(row[ri])
        name = text(row[ni]) if ni is not None and ni < len(row) else ""
        for j in score_cols:
            if j >= len(row):
                continue
            score = number(row[j])
            if pd.isna(score):
                continue
            ctx = [ffilled[k][j] for k in range(rubric_row) if j < len(ffilled[k]) and ffilled[k][j]]
            # Keep order while removing duplicates/noise.
            clean = []
            for v in ctx:
                if v not in clean and "campus name" not in normal(v) and "course name" not in normal(v) and "mentor name" not in normal(v):
                    clean.append(v)
            if label == "Practical Stock Market":
                week = next((v for v in clean if "week" in normal(v)), "")
                session = next((v for v in clean if "session" in normal(v)), "")
                activity = next((v for v in reversed(clean) if "session" not in normal(v) and "week" not in normal(v)), "")
                group = week
            else:
                dataset = next((v for v in clean if "dataset" in normal(v)), "")
                week = next((v for v in clean if "week" in normal(v)), "")
                session = next((v for v in clean if "session" in normal(v)), "")
                activity = next((v for v in reversed(clean) if not any(x in normal(v) for x in ["dataset", "week", "session"])), "")
                group = dataset
            sessions.append({
                "Register No": reg, "Student Name": name,
                "Group": group, "Week": week, "Session": session,
                "Activity": activity, "Score / 5": score,
            })
        for j in sorted(report_cols):
            if j >= len(row) or not text(row[j]):
                continue
            ctx = [ffilled[k][j] for k in range(rubric_row) if j < len(ffilled[k]) and ffilled[k][j]]
            reports.append({
                "Register No": reg, "Student Name": name,
                "Context": " • ".join(dict.fromkeys(v for v in ctx if v and "report link" not in normal(v))),
                "Report Link / File": text(row[j]),
            })
    return pd.DataFrame(sessions), pd.DataFrame(reports)


def parse_dsa_detail(wb, candidates, label):
    s = match_sheet(wb, candidates)
    if not s:
        return pd.DataFrame(), pd.DataFrame()
    rows = sheet_rows(wb, s)
    h = find_header(rows, ["register", "student", "questions completed"], 10)
    if h is None:
        return pd.DataFrame(), pd.DataFrame()
    hd = rows[h]
    ri = find_col(hd, ["register"])
    ni = find_col(hd, ["student", "name"])
    qi = find_col(hd, ["questions", "completed"])
    profile_i = find_col(hd, ["profile", "link"])
    if ri is None:
        return pd.DataFrame(), pd.DataFrame()

    metric_row, best = None, -1
    for i in range(h, min(h + 8, len(rows))):
        count = sum("question status" in normal(x) for x in rows[i])
        if count > best:
            metric_row, best = i, count
    if metric_row is None or best <= 0:
        return pd.DataFrame(), pd.DataFrame()
    data_start = first_student_row(rows, ri, metric_row + 1)
    if data_start is None:
        return pd.DataFrame(), pd.DataFrame()

    width = max(len(r) for r in rows[:metric_row + 1])
    top_ffill = ffill_values(rows[h], width)
    questions, summary = [], []
    status_cols = [j for j, v in enumerate(rows[metric_row]) if "question status" in normal(v)]
    total_assigned = len(status_cols)

    for row in rows[data_start:]:
        if ri >= len(row) or not register_like(row[ri]):
            continue
        reg = text(row[ri])
        name = text(row[ni]) if ni is not None and ni < len(row) else ""
        summary.append({
            "Register No": reg,
            "Student Name": name,
            "Questions Completed": number(row[qi]) if qi is not None and qi < len(row) else np.nan,
            "Total Questions Assigned": total_assigned,
            "HackerRank Profile": text(row[profile_i]) if profile_i is not None and profile_i < len(row) else "",
        })
        for j in status_cols:
            status = text(row[j]) if j < len(row) else ""
            # Keep tracked questions even when a mentor has not yet entered a status.
            if not status:
                status = "Not Assessed"
            question = ""
            link = ""
            date_context = ""
            for k in range(metric_row - 1, h, -1):
                if j >= len(rows[k]):
                    continue
                v = pretty_header(rows[k][j])
                if not v:
                    continue
                if v.startswith("http"):
                    link = link or v
                elif not question and not re.fullmatch(r"week[s]?\s*\d+", normal(v)):
                    question = v
                if re.search(r"\d{1,2}[/\-]\d{1,2}/?\d{0,4}", v):
                    date_context = v
            week = top_ffill[j] if j < len(top_ffill) and "week" in normal(top_ffill[j]) else ""
            mark = np.nan
            if j + 1 < len(rows[metric_row]) and "class work" in normal(rows[metric_row][j + 1]) and j + 1 < len(row):
                mark = number(row[j + 1])
            questions.append({
                "Register No": reg, "Student Name": name,
                "Week": week, "Question / Topic": question,
                "Date / Context": date_context, "Status": status,
                "Class Work / 5": mark, "Question Link": link,
            })
    return pd.DataFrame(summary), pd.DataFrame(questions)

# -------------------- DATA SOURCES --------------------
def needed_google_candidates(programme):
    common = [
        ["Attendance"],
        ["Aptitude Assessment", "Aptitude  Assessment"],
        ["Aptitude"],
        ["Communication"],
    ]
    if programme == "B.Com":
        return common + [
            ["Practical Stock Market Analysis"],
            ["Financial Data Analysis using P", "Financial Data Analysis"],
        ]
    return common + [
        ["DSA -Skill Development", "DSA Skill Development"],
        ["DSA Extra Questions Skill", "DSA Extra Questions"],
    ]


def _match_title(titles, candidates):
    for c in candidates:
        for t in titles:
            if normal(c) == normal(t):
                return t
    for c in candidates:
        for t in titles:
            if normal(c) in normal(t):
                return t
    return None


def google_secrets_available():
    try:
        return (
            "google_service_account" in st.secrets
            and "google_sheets" in st.secrets
            and bool(st.secrets["google_sheets"].get("bcom_spreadsheet_id", ""))
            and bool(st.secrets["google_sheets"].get("bca_spreadsheet_id", ""))
        )
    except Exception:
        return False


@st.cache_resource
def google_client():
    import gspread
    from google.oauth2.service_account import Credentials

    info = dict(st.secrets["google_service_account"])
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets.readonly",
        "https://www.googleapis.com/auth/drive.readonly",
    ]
    creds = Credentials.from_service_account_info(info, scopes=scopes)
    return gspread.authorize(creds)


@st.cache_data(ttl=REFRESH_SECONDS, show_spinner="Loading latest Google Sheets data...")
def read_google_tracker(spreadsheet_id, programme):
    gc = google_client()
    sh = gc.open_by_key(spreadsheet_id)
    worksheets = sh.worksheets()
    by_title = {w.title: w for w in worksheets}
    titles = list(by_title)

    sheet_map = {}
    for candidates in needed_google_candidates(programme):
        title = _match_title(titles, candidates)
        if title and title not in sheet_map:
            sheet_map[title] = by_title[title].get_all_values()

    wb = GoogleRowsBook(sheet_map)
    return parse_tracker_book(wb, programme, f"Google Sheets • {sh.title}")


@st.cache_data(show_spinner="Reading local tracker data...")
def read_local_tracker(path, mtime, programme):
    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    return parse_tracker_book(wb, programme, Path(path).name)


@st.cache_data(show_spinner="Reading uploaded tracker data...")
def read_uploaded_tracker(file_bytes, filename, programme):
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True, keep_links=False)
    return parse_tracker_book(wb, programme, filename)


def parse_tracker_book(wb, programme, source_name):
    attendance = parse_attendance(wb)
    attendance_daily = parse_attendance_daily(wb)
    aptitude = parse_aptitude_assessment(wb)
    aptitude_tasks = parse_aptitude_tasks(wb)
    aptitude_summary_detail, aptitude_ca_detail = parse_aptitude_assessment_detail(wb)
    aptitude_task_detail = parse_aptitude_task_detail(wb)
    communication = parse_communication(wb)
    communication_detail = parse_communication_detail(wb)

    practical = pd.DataFrame()
    financial = pd.DataFrame()
    dsa_core = pd.DataFrame()
    dsa_extra = pd.DataFrame()
    practical_detail = pd.DataFrame()
    practical_reports = pd.DataFrame()
    financial_detail = pd.DataFrame()
    financial_reports = pd.DataFrame()
    dsa_core_summary = pd.DataFrame()
    dsa_core_detail = pd.DataFrame()
    dsa_extra_summary = pd.DataFrame()
    dsa_extra_detail = pd.DataFrame()

    if programme == "B.Com":
        practical = parse_outof5_skill(
            wb, ["Practical Stock Market Analysis"], "Practical Stock Market"
        )
        financial = parse_outof5_skill(
            wb, ["Financial Data Analysis using P", "Financial Data Analysis"], "Financial Data Analytics"
        )
        practical_detail, practical_reports = parse_bcom_skill_detail(
            wb, ["Practical Stock Market Analysis"], "Practical Stock Market"
        )
        financial_detail, financial_reports = parse_bcom_skill_detail(
            wb, ["Financial Data Analysis using P", "Financial Data Analysis"], "Financial Data Analytics"
        )
    else:
        dsa_core = parse_dsa_sheet(
            wb, ["DSA -Skill Development", "DSA Skill Development"], "DSA Skill Development"
        )
        dsa_extra = parse_dsa_sheet(
            wb, ["DSA Extra Questions Skill", "DSA Extra Questions"], "DSA Extra Questions Skill"
        )
        dsa_core_summary, dsa_core_detail = parse_dsa_detail(
            wb, ["DSA -Skill Development", "DSA Skill Development"], "DSA Skill Development"
        )
        dsa_extra_summary, dsa_extra_detail = parse_dsa_detail(
            wb, ["DSA Extra Questions Skill", "DSA Extra Questions"], "DSA Extra Questions Skill"
        )

    master = attendance.copy()
    for d in [aptitude, aptitude_tasks, communication, practical, financial, dsa_core, dsa_extra]:
        master = merge_frames(master, d)

    if master is None or master.empty:
        return {"programme": programme, "source_name": source_name, "master": pd.DataFrame()}

    if "Student Name" not in master.columns:
        master["Student Name"] = ""

    apt_cols = [c for c in ["Aptitude Accuracy %", "Aptitude Task Score %"] if c in master.columns]
    master["Aptitude Score %"] = master[apt_cols].mean(axis=1, skipna=True) if apt_cols else np.nan

    # DSA progress is based on completed questions / total questions assigned.
    # It is NOT normalized against the best-performing student.
    if programme == "BCA":
        dsa_specs = [
            ("DSA Skill Development", "DSA Skill Development %", dsa_core_summary),
            ("DSA Extra Questions Skill", "DSA Extra Questions Skill %", dsa_extra_summary),
        ]
        for raw, normalized, summary in dsa_specs:
            assigned_col = f"{raw} Assigned"
            if not summary.empty and "Total Questions Assigned" in summary.columns:
                assigned_map = summary.drop_duplicates("Register No").set_index("Register No")["Total Questions Assigned"]
                master[assigned_col] = master["Register No"].astype(str).map(assigned_map)
                denom = pd.to_numeric(master[assigned_col], errors="coerce")
                numer = pd.to_numeric(master.get(raw), errors="coerce") if raw in master.columns else pd.Series(np.nan, index=master.index)
                master[normalized] = np.where(denom > 0, (numer / denom * 100).clip(0, 100), np.nan)
            else:
                master[assigned_col] = np.nan
                master[normalized] = np.nan

    if programme == "B.Com":
        component_cols = [
            "Attendance %",
            "Communication Score %",
            "Aptitude Score %",
            "Practical Stock Market Score %",
            "Financial Data Analytics Score %",
        ]
    else:
        component_cols = [
            "Attendance %",
            "Communication Score %",
            "Aptitude Score %",
            "DSA Skill Development %",
            "DSA Extra Questions Skill %",
        ]

    available = [c for c in component_cols if c in master.columns]
    counts = master[available].notna().sum(axis=1)
    master["Composite Performance %"] = master[available].mean(axis=1, skipna=True)
    master.loc[counts < COMPOSITE_MIN_COMPONENTS, "Composite Performance %"] = np.nan
    master["Assessed Components"] = counts

    def classify(row):
        att = row.get("Attendance %", np.nan)
        skill_vals = [row.get(c, np.nan) for c in component_cols if c != "Attendance %"]
        skill_vals = [float(v) for v in skill_vals if pd.notna(v)]
        critical_skills = sum(v < SKILL_CRITICAL for v in skill_vals)
        attention_skills = sum(v < SKILL_ON_TRACK for v in skill_vals)
        if (pd.notna(att) and att < ATTENDANCE_AT_RISK) or critical_skills >= 2:
            return "At Risk"
        if (pd.notna(att) and att < ATTENDANCE_ON_TRACK) or attention_skills >= 1:
            return "Needs Attention"
        return "On Track"

    master["Performance Status"] = master.apply(classify, axis=1)
    master = master.drop_duplicates(subset=["Register No"], keep="first")
    master = master.sort_values(
        ["Composite Performance %", "Student Name"], ascending=[False, True], na_position="last"
    ).reset_index(drop=True)

    return {
        "programme": programme,
        "source_name": source_name,
        "master": master,
        "attendance": attendance,
        "attendance_daily": attendance_daily,
        "aptitude": aptitude,
        "aptitude_summary_detail": aptitude_summary_detail,
        "aptitude_ca_detail": aptitude_ca_detail,
        "aptitude_task_detail": aptitude_task_detail,
        "communication": communication,
        "communication_detail": communication_detail,
        "practical": practical,
        "practical_detail": practical_detail,
        "practical_reports": practical_reports,
        "financial": financial,
        "financial_detail": financial_detail,
        "financial_reports": financial_reports,
        "dsa_core": dsa_core,
        "dsa_core_summary": dsa_core_summary,
        "dsa_core_detail": dsa_core_detail,
        "dsa_extra": dsa_extra,
        "dsa_extra_summary": dsa_extra_summary,
        "dsa_extra_detail": dsa_extra_detail,
        "component_cols": component_cols,
    }


def component_label_map(programme):
    return {field: label for label, field in percent_specs(programme)}


def metric_status(value, is_attendance=False):
    if pd.isna(value):
        return "Not Assessed"
    if is_attendance:
        if value < ATTENDANCE_AT_RISK:
            return "At Risk"
        if value < ATTENDANCE_ON_TRACK:
            return "Needs Attention"
        return "On Track"
    if value < SKILL_CRITICAL:
        return "Critical"
    if value < SKILL_ON_TRACK:
        return "Needs Attention"
    return "On Track"


def performance_reasons(row, programme):
    reasons = []
    att = row.get("Attendance %", np.nan)
    if pd.notna(att):
        if att < ATTENDANCE_AT_RISK:
            reasons.append(f"Attendance {att:.1f}% is below the {ATTENDANCE_AT_RISK:.0f}% At Risk benchmark.")
        elif att < ATTENDANCE_ON_TRACK:
            reasons.append(f"Attendance {att:.1f}% is within the {ATTENDANCE_AT_RISK:.0f}-{ATTENDANCE_ON_TRACK - 0.1:.1f}% monitoring band.")

    critical = []
    attention = []
    for label, field in percent_specs(programme):
        if field == "Attendance %":
            continue
        v = row.get(field, np.nan)
        if pd.isna(v):
            continue
        if v < SKILL_CRITICAL:
            critical.append(f"{label} ({v:.1f}%)")
        elif v < SKILL_ON_TRACK:
            attention.append(f"{label} ({v:.1f}%)")
    if len(critical) >= 2:
        reasons.append("Two or more skill areas are below the 40% critical benchmark: " + ", ".join(critical) + ".")
    elif critical:
        reasons.append("Critical skill area: " + ", ".join(critical) + ".")
    if attention:
        reasons.append("Skill areas below the 60% On Track benchmark: " + ", ".join(attention) + ".")
    if not reasons:
        reasons.append("Attendance is at least 80% and all assessed skill areas are at or above 60%.")
    return reasons


def performance_alerts(row, programme):
    alerts = []
    att = row.get("Attendance %", np.nan)
    if pd.notna(att) and att < ATTENDANCE_AT_RISK:
        alerts.append(("risk", "Attendance", f"{att:.1f}% - below {ATTENDANCE_AT_RISK:.0f}% At Risk benchmark"))
    elif pd.notna(att) and att < ATTENDANCE_ON_TRACK:
        alerts.append(("attention", "Attendance", f"{att:.1f}% - monitor until attendance reaches {ATTENDANCE_ON_TRACK:.0f}%"))

    for label, field in percent_specs(programme):
        if field == "Attendance %":
            continue
        v = row.get(field, np.nan)
        if pd.isna(v):
            continue
        if v < SKILL_CRITICAL:
            alerts.append(("risk", label, f"{v:.1f}% - below {SKILL_CRITICAL:.0f}% critical benchmark"))
        elif v < SKILL_ON_TRACK:
            alerts.append(("attention", label, f"{v:.1f}% - below {SKILL_ON_TRACK:.0f}% On Track benchmark"))
    return alerts


def comparison_data(row, master, programme):
    rows = []
    for label, field in percent_specs(programme):
        sv = row.get(field, np.nan)
        cv = master[field].mean(skipna=True) if field in master.columns else np.nan
        rows.append({
            "Dimension": label,
            "Student %": sv,
            "Class Average %": cv,
            "Difference": sv - cv if pd.notna(sv) and pd.notna(cv) else np.nan,
            "Status": metric_status(sv, is_attendance=(field == "Attendance %")),
        })
    return pd.DataFrame(rows)


def benchmark_comparison(label, field, student_row, master, color, is_attendance=False):
    student = student_row.get(field, np.nan)
    class_avg = master[field].mean(skipna=True) if field in master.columns else np.nan
    diff = student - class_avg if pd.notna(student) and pd.notna(class_avg) else np.nan
    st.markdown(f"#### {label} - Student vs Class Benchmark")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Student", f"{student:.1f}%" if pd.notna(student) else "N/A")
    c2.metric("Class Average", f"{class_avg:.1f}%" if pd.notna(class_avg) else "N/A")
    c3.metric("Difference", f"{diff:+.1f}%" if pd.notna(diff) else "N/A")
    c4.metric("Monitoring Status", metric_status(student, is_attendance=is_attendance))
    if pd.notna(student) and pd.notna(class_avg):
        x = pd.DataFrame({"Comparison": ["Student", "Class Average"], "Score": [student, class_avg]})
        fig = px.bar(x, x="Score", y="Comparison", orientation="h", text_auto=".1f", color_discrete_sequence=[color])
        fig.update_xaxes(range=[0, 100], title="Score / Progress %")
        fig.update_layout(showlegend=False)
        st.plotly_chart(plot_layout(fig, 210), use_container_width=True)


def render_alert_panel(row, programme):
    alerts = performance_alerts(row, programme)
    st.markdown("#### Performance Alerts")
    if not alerts:
        st.markdown('<div class="alert-ok"><b>No critical performance alerts.</b><br>Current assessed indicators meet the dashboard monitoring benchmarks.</div>', unsafe_allow_html=True)
        return
    risk_count = sum(level == "risk" for level, _, _ in alerts)
    attention_count = sum(level == "attention" for level, _, _ in alerts)
    if risk_count:
        st.markdown(f'<div class="alert-risk"><b>{risk_count} critical alert(s)</b> and {attention_count} monitoring alert(s) require mentor review.</div>', unsafe_allow_html=True)
    else:
        st.markdown(f'<div class="alert-attention"><b>{attention_count} area(s) need attention.</b> No current critical alert is triggered.</div>', unsafe_allow_html=True)
    for level, label, message in alerts:
        css = "alert-risk" if level == "risk" else "alert-attention"
        st.markdown(f'<div class="{css}"><b>{label}</b><br>{message}</div>', unsafe_allow_html=True)


def benchmark_explainer(programme):
    with st.expander("How are performance status and composite score calculated?"):
        st.markdown("**Dashboard Monitoring Benchmark**  \nThis is a configurable dashboard rule, not an official benchmark stored in Google Sheets.")
        a, b = st.columns(2)
        with a:
            st.markdown(
                f"**Attendance**  \n"
                f"- **At Risk:** below {ATTENDANCE_AT_RISK:.0f}%  \n"
                f"- **Needs Attention:** {ATTENDANCE_AT_RISK:.0f}% to {ATTENDANCE_ON_TRACK - 0.1:.1f}%  \n"
                f"- **On Track:** {ATTENDANCE_ON_TRACK:.0f}% and above"
            )
        with b:
            st.markdown(
                f"**Skill performance**  \n"
                f"- **Critical:** below {SKILL_CRITICAL:.0f}%  \n"
                f"- **Needs Attention:** {SKILL_CRITICAL:.0f}% to {SKILL_ON_TRACK - 0.1:.1f}%  \n"
                f"- **On Track:** {SKILL_ON_TRACK:.0f}% and above"
            )
        st.markdown(
            f"**Overall status**  \n"
            f"- **At Risk:** attendance below {ATTENDANCE_AT_RISK:.0f}% **or** at least two assessed skill areas below {SKILL_CRITICAL:.0f}%.  \n"
            f"- **Needs Attention:** not At Risk, but attendance below {ATTENDANCE_ON_TRACK:.0f}% **or** any assessed skill area below {SKILL_ON_TRACK:.0f}%.  \n"
            f"- **On Track:** attendance at least {ATTENDANCE_ON_TRACK:.0f}% and all assessed skill areas at least {SKILL_ON_TRACK:.0f}%."
        )
        weights = pd.DataFrame({"Element": [x[0] for x in percent_specs(programme)], "Weight when all 5 are assessed": ["20%"] * 5})
        st.markdown("**Composite Performance**")
        st.dataframe(weights, hide_index=True, use_container_width=True)
        st.caption("The composite is the equal-weight mean of assessed elements. When all five elements are assessed, each contributes 20%. Missing/unassessed values are not treated as zero; a composite is shown only when at least three elements are available.")
        if programme == "BCA":
            st.info("DSA progress is calculated as Questions Completed / Total Questions Assigned × 100. It is not normalized against the top-performing student.")


def _pdf_safe(v, max_len=90):
    s = text(v).replace("\n", " ")
    return s if len(s) <= max_len else s[:max_len - 1] + "…"


def build_student_pdf(student_row, master, data, programme, source_mode):
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT, TA_CENTER
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=14*mm, leftMargin=14*mm, topMargin=14*mm, bottomMargin=14*mm)
    styles = getSampleStyleSheet()
    title = ParagraphStyle('TitleX', parent=styles['Title'], fontName='Helvetica-Bold', fontSize=18, leading=22, textColor=colors.HexColor(TEXT), alignment=TA_LEFT)
    h2 = ParagraphStyle('H2X', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=12, leading=15, textColor=colors.HexColor(TEAL_DARK), spaceBefore=8, spaceAfter=6)
    body = ParagraphStyle('BodyX', parent=styles['BodyText'], fontSize=8.5, leading=11, textColor=colors.HexColor(TEXT))
    small = ParagraphStyle('SmallX', parent=body, fontSize=7.5, leading=9)
    story = []

    name = text(student_row.get('Student Name', ''))
    reg = text(student_row.get('Register No', ''))
    status = text(student_row.get('Performance Status', ''))
    story.append(Paragraph("Student Performance Report", title))
    story.append(Spacer(1, 3*mm))
    source_label = "Live Google Sheets" if source_mode == "Live Google Sheets" else _pdf_safe(data.get('source_name',''), 48)
    meta = [
        ["Student", Paragraph(_pdf_safe(name, 45), small), "Register No.", Paragraph(_pdf_safe(reg, 32), small)],
        ["Programme", Paragraph(_pdf_safe(f"{programme} - {SEMESTER}", 45), small), "Academic Year", Paragraph(_pdf_safe(ACADEMIC_YEAR, 20), small)],
        ["Performance Status", Paragraph(_pdf_safe(status, 30), small), "Composite Performance", Paragraph(display_value(student_row.get('Composite Performance %', np.nan), 'percent'), small)],
        ["Data Source", Paragraph(source_label, small), "Generated", Paragraph(datetime.now(ZoneInfo("Asia/Kolkata")).strftime('%d %b %Y, %I:%M %p'), small)],
    ]
    t = Table(meta, colWidths=[30*mm, 58*mm, 35*mm, 55*mm])
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(0,-1),colors.HexColor('#EFF8F6')), ('BACKGROUND',(2,0),(2,-1),colors.HexColor('#EFF8F6')),
        ('TEXTCOLOR',(0,0),(-1,-1),colors.HexColor(TEXT)), ('FONTNAME',(0,0),(0,-1),'Helvetica-Bold'), ('FONTNAME',(2,0),(2,-1),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,-1),8), ('GRID',(0,0),(-1,-1),0.35,colors.HexColor(BORDER)), ('VALIGN',(0,0),(-1,-1),'MIDDLE'), ('PADDING',(0,0),(-1,-1),5),
    ]))
    story.append(t)

    story.append(Paragraph("Five-Element Benchmark", h2))
    comp = comparison_data(student_row, master, programme)
    comp_rows = [["Dimension","Student","Class Avg","Difference","Status"]]
    for _, r in comp.iterrows():
        comp_rows.append([
            _pdf_safe(r['Dimension'], 28),
            f"{r['Student %']:.1f}%" if pd.notna(r['Student %']) else "N/A",
            f"{r['Class Average %']:.1f}%" if pd.notna(r['Class Average %']) else "N/A",
            f"{r['Difference']:+.1f}%" if pd.notna(r['Difference']) else "N/A",
            _pdf_safe(r['Status'], 22),
        ])
    ct = Table(comp_rows, repeatRows=1, colWidths=[45*mm,28*mm,28*mm,28*mm,38*mm])
    ct.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor(TEAL_DARK)), ('TEXTCOLOR',(0,0),(-1,0),colors.white), ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,-1),7.8), ('GRID',(0,0),(-1,-1),0.35,colors.HexColor(BORDER)), ('PADDING',(0,0),(-1,-1),4), ('VALIGN',(0,0),(-1,-1),'MIDDLE')
    ]))
    story.append(ct)

    story.append(Paragraph("Performance Alerts and Status Rationale", h2))
    alerts = performance_alerts(student_row, programme)
    reasons = performance_reasons(student_row, programme)
    if alerts:
        for level, label, msg in alerts:
            prefix = "CRITICAL" if level == "risk" else "ATTENTION"
            story.append(Paragraph(f"<b>{prefix} - {label}:</b> {_pdf_safe(msg, 160)}", body))
    else:
        story.append(Paragraph("No critical performance alerts under the current dashboard monitoring benchmark.", body))
    story.append(Spacer(1, 2*mm))
    for reason in reasons:
        story.append(Paragraph("- " + _pdf_safe(reason, 180), body))

    # Detailed activity information
    story.append(Paragraph("Detailed Performance Information", h2))
    att = data.get('attendance', pd.DataFrame())
    ar = att[att['Register No'].astype(str) == reg] if not att.empty else pd.DataFrame()
    if not ar.empty:
        a = ar.iloc[0]
        rows = [["Attendance", "Value"], ["Total Days", str(int(a['Total Days'])) if pd.notna(a.get('Total Days')) else 'N/A'], ["Present Days", str(int(a['Present Days'])) if pd.notna(a.get('Present Days')) else 'N/A'], ["Absent Days", str(int(a['Absent Days'])) if pd.notna(a.get('Absent Days')) else 'N/A'], ["Attendance %", display_value(a.get('Attendance %', np.nan),'percent')]]
        at = Table(rows, colWidths=[50*mm,45*mm])
        at.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#EFF8F6')),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),8),('GRID',(0,0),(-1,-1),0.35,colors.HexColor(BORDER)),('PADDING',(0,0),(-1,-1),4)]))
        story.append(at)

    def add_detail(title_text, frame, columns, max_rows=None):
        if frame is None or frame.empty:
            return
        f = frame[frame['Register No'].astype(str) == reg].copy() if 'Register No' in frame.columns else frame.copy()
        if f.empty:
            return
        if max_rows:
            f = f.head(max_rows)
        usable = [c for c in columns if c in f.columns]
        if not usable:
            return
        story.append(Paragraph(title_text, h2))
        table_rows = [[Paragraph(str(c), small) for c in usable]]
        for _, rr in f[usable].iterrows():
            table_rows.append([Paragraph(_pdf_safe(rr[c], 80), small) for c in usable])
        widths = [max(22*mm, 175*mm/len(usable))] * len(usable)
        widths[-1] += 175*mm - sum(widths)
        tb = Table(table_rows, repeatRows=1, colWidths=widths)
        tb.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#EFF8F6')),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('GRID',(0,0),(-1,-1),0.25,colors.HexColor(BORDER)),('VALIGN',(0,0),(-1,-1),'TOP'),('PADDING',(0,0),(-1,-1),3)]))
        story.append(tb)

    add_detail("Communication Activity Details", data.get('communication_detail', pd.DataFrame()), ["Activity","Date / Context","Criteria Breakdown","Total","CEFR"])
    add_detail("Aptitude Continuous Assessments", data.get('aptitude_ca_detail', pd.DataFrame()), ["Assessment","Date","Questions Solved","Correct Answers","Marks %","Assessment Status"])
    add_detail("Aptitude Practice / Task Details", data.get('aptitude_task_detail', pd.DataFrame()), ["Date","Task / Assessment","Topic","Category","Score","Max Score","Status"])
    if programme == "B.Com":
        add_detail("Practical Stock Market - Session Details", data.get('practical_detail', pd.DataFrame()), ["Group","Week","Session","Activity","Score / 5"])
        add_detail("Financial Data Analytics - Session Details", data.get('financial_detail', pd.DataFrame()), ["Group","Week","Session","Activity","Score / 5"])
    else:
        add_detail("DSA Skill Development - Question Details", data.get('dsa_core_detail', pd.DataFrame()), ["Week","Question / Topic","Date / Context","Status","Class Work / 5"])
        add_detail("DSA Extra Questions - Question Details", data.get('dsa_extra_detail', pd.DataFrame()), ["Week","Question / Topic","Date / Context","Status","Class Work / 5"])

    story.append(Spacer(1, 4*mm))
    story.append(Paragraph("Dashboard Monitoring Benchmark", h2))
    story.append(Paragraph(
        f"Attendance: At Risk below {ATTENDANCE_AT_RISK:.0f}%; Needs Attention {ATTENDANCE_AT_RISK:.0f}-{ATTENDANCE_ON_TRACK-0.1:.1f}%; On Track at least {ATTENDANCE_ON_TRACK:.0f}%. "
        f"Skills: Critical below {SKILL_CRITICAL:.0f}%; Needs Attention {SKILL_CRITICAL:.0f}-{SKILL_ON_TRACK-0.1:.1f}%; On Track at least {SKILL_ON_TRACK:.0f}%. "
        "This benchmark is configured in the dashboard and is not an official benchmark stored in Google Sheets.", body))
    doc.build(story)
    return buf.getvalue()


# -------------------- DISPLAY HELPERS --------------------
def metric_specs(programme):
    if programme == "B.Com":
        return [
            ("Attendance", "Attendance %", "percent", TEAL),
            ("Communication", "Communication Score %", "percent", YELLOW),
            ("Aptitude", "Aptitude Score %", "percent", PINK),
            ("Practical Stock Market", "Practical Stock Market Score %", "percent", BLUE),
            ("Financial Data Analytics", "Financial Data Analytics Score %", "percent", PURPLE),
        ]
    return [
        ("Attendance", "Attendance %", "percent", TEAL),
        ("Communication", "Communication Score %", "percent", YELLOW),
        ("Aptitude", "Aptitude Score %", "percent", PINK),
        ("DSA Skill Development", "DSA Skill Development", "count", BLUE),
        ("DSA Extra Questions Skill", "DSA Extra Questions Skill", "count", PURPLE),
    ]


def percent_specs(programme):
    if programme == "B.Com":
        return [
            ("Attendance", "Attendance %"),
            ("Communication", "Communication Score %"),
            ("Aptitude", "Aptitude Score %"),
            ("Practical Stock Market", "Practical Stock Market Score %"),
            ("Financial Data Analytics", "Financial Data Analytics Score %"),
        ]
    return [
        ("Attendance", "Attendance %"),
        ("Communication", "Communication Score %"),
        ("Aptitude", "Aptitude Score %"),
        ("DSA Skill Development", "DSA Skill Development %"),
        ("DSA Extra Questions", "DSA Extra Questions Skill %"),
    ]


def display_value(v, kind):
    if pd.isna(v):
        return "N/A"
    if kind == "count":
        return f"{v:,.1f}"
    return f"{v:,.1f}%"


def card_html(label, value, kind, color, note="Class / filtered average"):
    return f"""
    <div class="metric-card">
      <div class="accent-line" style="background:{color};"></div>
      <div class="metric-label">{label}</div>
      <div class="metric-value">{display_value(value, kind)}</div>
      <div class="metric-note">{note}</div>
    </div>
    """


def status_color(status):
    return {"On Track": TEAL_DARK, "Needs Attention": YELLOW, "At Risk": PINK}.get(status, MUTED)


def apply_filters(master):
    with st.sidebar:
        st.markdown("### Filters")
        status = st.selectbox("Performance Status", ["All", "On Track", "Needs Attention", "At Risk"], key="filter_status")

        options_df = master[["Register No", "Student Name"]].drop_duplicates().sort_values("Student Name")
        option_map = {"All Students": None}
        for _, r in options_df.iterrows():
            label = f"{r['Student Name']}  •  {r['Register No']}"
            option_map[label] = str(r["Register No"])
        selected_student = st.selectbox("Student", list(option_map.keys()), key="filter_student")

        attendance_band = st.selectbox(
            "Attendance",
            [
                "All",
                f"At Risk (<{ATTENDANCE_AT_RISK:.0f}%)",
                f"Needs Attention ({ATTENDANCE_AT_RISK:.0f}-{ATTENDANCE_ON_TRACK - 0.1:.1f}%)",
                f"On Track (≥{ATTENDANCE_ON_TRACK:.0f}%)",
            ],
            key="filter_attendance",
        )

        if st.button("Reset filters", use_container_width=True):
            for key in ["filter_status", "filter_student", "filter_attendance"]:
                st.session_state.pop(key, None)
            st.rerun()

    filtered = master.copy()
    if status != "All":
        filtered = filtered[filtered["Performance Status"] == status]
    reg = option_map[selected_student]
    if reg:
        filtered = filtered[filtered["Register No"].astype(str) == reg]
    if "Attendance %" in filtered.columns:
        if attendance_band.startswith("At Risk"):
            filtered = filtered[filtered["Attendance %"] < ATTENDANCE_AT_RISK]
        elif attendance_band.startswith("Needs Attention"):
            filtered = filtered[(filtered["Attendance %"] >= ATTENDANCE_AT_RISK) & (filtered["Attendance %"] < ATTENDANCE_ON_TRACK)]
        elif attendance_band.startswith("On Track"):
            filtered = filtered[filtered["Attendance %"] >= ATTENDANCE_ON_TRACK]
    active_filters = status != "All" or attendance_band != "All"
    return filtered, reg, active_filters

def hero(programme_label, data, filtered, source_mode, connection_ok=True):
    refreshed = datetime.now(ZoneInfo("Asia/Kolkata")).strftime("%d %b %Y, %I:%M %p")
    total = len(data["master"])
    shown = len(filtered)
    source_text = "Live Google Sheets" if source_mode == "Live Google Sheets" else data["source_name"]
    live_html = ""
    if source_mode == "Live Google Sheets":
        if connection_ok:
            live_html = f'<div style="margin-top:12px;"><span class="live-chip"><span class="live-dot"></span>LIVE · Google Sheets connected · Auto-refresh {REFRESH_SECONDS}s</span></div>'
        else:
            live_html = '<div style="margin-top:12px;"><span class="connection-bad">DATA CONNECTION ISSUE · Showing last available data</span></div>'
    st.markdown(
        f"""
        <div class="hero">
          <h1>📊 {APP_TITLE}</h1>
          <p>{APP_SUBTITLE}</p>
          <div class="hero-meta">
            <b>{programme_label}</b> &nbsp; • &nbsp; {ACADEMIC_YEAR}
            &nbsp; • &nbsp; Showing {shown} of {total} students<br>
            Source: {source_text} &nbsp; • &nbsp; Last synced: {refreshed}
          </div>
          {live_html}
        </div>
        """,
        unsafe_allow_html=True,
    )

def top_cards(filtered, master, programme, selected_reg, active_filters=False):
    cols = st.columns(5)
    if selected_reg:
        note = "Student Performance"
    elif active_filters or len(filtered) != len(master):
        note = "Filtered Class Average"
    else:
        note = "Class Average"
    for col, (label, field, kind, color) in zip(cols, metric_specs(programme)):
        v = filtered[field].mean(skipna=True) if field in filtered.columns and not filtered.empty else np.nan
        if kind == "count" and not selected_reg:
            note_this = "Avg questions completed" if not active_filters else "Filtered avg questions"
        elif kind == "count":
            note_this = "Questions Completed"
        else:
            note_this = note
        with col:
            st.markdown(card_html(label, v, kind, color, note_this), unsafe_allow_html=True)

def plot_layout(fig, height=360):
    fig.update_layout(
        height=height,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=10, r=10, t=25, b=10),
        font=dict(color=TEXT),
    )
    return fig


def overview_tab(filtered, programme, data=None, selected_reg=None, source_mode="Local Excel"):
    master = data["master"] if data is not None else filtered
    if selected_reg:
        student = master[master["Register No"].astype(str) == str(selected_reg)]
        if student.empty:
            st.info("Student overview is not available.")
            return
        r = student.iloc[0]
        status = r.get("Performance Status", "")
        sc = status_color(status)
        st.markdown(
            f'''<div class="student-banner">
                <div style="display:flex;justify-content:space-between;gap:14px;align-items:flex-start;flex-wrap:wrap;">
                  <div><div style="font-size:1.5rem;font-weight:800;color:{TEXT};">{r.get('Student Name','')} - Student 360° Overview</div>
                  <div style="color:{MUTED};margin-top:4px;">Register No: {r.get('Register No','')} · {programme} · {SEMESTER}</div></div>
                  <span class="status-pill" style="background:{sc}22;color:{sc};font-size:.88rem;">{status.upper()}</span>
                </div>
              </div>''', unsafe_allow_html=True)

        c1, c2 = st.columns([1.2, .8])
        with c1:
            comp = comparison_data(r, master, programme)
            st.markdown("#### Student vs Class Benchmark")
            view = comp.copy()
            for c in ["Student %", "Class Average %", "Difference"]:
                view[c] = view[c].map(lambda x: f"{x:+.1f}%" if c == "Difference" and pd.notna(x) else (f"{x:.1f}%" if pd.notna(x) else "N/A"))
            st.dataframe(view, use_container_width=True, hide_index=True)
        with c2:
            st.markdown("#### Five-Element Profile")
            cats, vals = [], []
            for label, field in percent_specs(programme):
                v = r.get(field, np.nan)
                if pd.notna(v):
                    cats.append(label); vals.append(v)
            if len(vals) >= 3:
                fig = go.Figure(go.Scatterpolar(r=vals + [vals[0]], theta=cats + [cats[0]], fill="toself", line=dict(color=TEAL_DARK), fillcolor="rgba(86,189,177,.20)"))
                fig.update_layout(polar=dict(radialaxis=dict(visible=True, range=[0,100])), showlegend=False)
                st.plotly_chart(plot_layout(fig, 350), use_container_width=True)

        render_alert_panel(r, programme)
        reasons = performance_reasons(r, programme)
        st.markdown("#### Why this status?")
        for reason in reasons:
            st.markdown(f"- {reason}")

        specs = percent_specs(programme)
        strengths = [label for label, field in specs if pd.notna(r.get(field, np.nan)) and r.get(field) >= STRENGTH_BENCHMARK]
        needs = [label for label, field in specs if pd.notna(r.get(field, np.nan)) and r.get(field) < SKILL_ON_TRACK and field != "Attendance %"]
        a,b = st.columns(2)
        with a:
            st.markdown('<div class="insight-card"><div class="insight-title">Strengths</div>' + ("<br>".join("• "+x for x in strengths) if strengths else "No dimension is currently at or above the 75% strength marker.") + '</div>', unsafe_allow_html=True)
        with b:
            st.markdown('<div class="insight-card"><div class="insight-title">Areas requiring mentor attention</div>' + ("<br>".join("• "+x for x in needs) if needs else "No assessed skill is currently below the 60% On Track benchmark.") + '</div>', unsafe_allow_html=True)

        benchmark_explainer(programme)
        try:
            pdf = build_student_pdf(r, master, data, programme, source_mode)
            st.download_button("Download Student Performance Report (PDF)", pdf, f"{text(r.get('Student Name','Student')).replace(' ','_')}_performance_report.pdf", "application/pdf", use_container_width=False, key=f"overview_pdf_{text(r.get('Register No','student'))}")
        except ModuleNotFoundError:
            st.warning("PDF report support requires ReportLab. Install it with: pip install reportlab")
        return

    left, middle, right = st.columns([1, 1.15, 1])
    with left:
        st.markdown('<div class="section-title">Performance Status</div>', unsafe_allow_html=True)
        counts = filtered["Performance Status"].value_counts().reindex(["On Track", "Needs Attention", "At Risk"], fill_value=0).reset_index()
        counts.columns = ["Status", "Students"]
        fig = px.bar(counts, x="Status", y="Students", text="Students", color="Status", color_discrete_map={"On Track": TEAL, "Needs Attention": YELLOW, "At Risk": PINK})
        fig.update_layout(showlegend=False)
        st.plotly_chart(plot_layout(fig, 330), use_container_width=True)
    with middle:
        st.markdown('<div class="section-title">Top Composite Performance</div>', unsafe_allow_html=True)
        x = filtered[["Student Name", "Composite Performance %"]].dropna().nlargest(10, "Composite Performance %")
        if x.empty:
            st.info("No composite score available for the current filter.")
        else:
            fig = px.bar(x.sort_values("Composite Performance %"), x="Composite Performance %", y="Student Name", orientation="h", text_auto=".1f", color_discrete_sequence=[BLUE])
            st.plotly_chart(plot_layout(fig, 330), use_container_width=True)
    with right:
        st.markdown('<div class="section-title">Five-Element Profile</div>', unsafe_allow_html=True)
        cats, vals = [], []
        for label, field in percent_specs(programme):
            if field in filtered.columns:
                v = filtered[field].mean(skipna=True)
                if pd.notna(v): cats.append(label); vals.append(v)
        if len(vals) >= 3:
            fig = go.Figure(go.Scatterpolar(r=vals+[vals[0]], theta=cats+[cats[0]], fill="toself", line=dict(color=TEAL_DARK), fillcolor="rgba(86,189,177,.20)"))
            fig.update_layout(polar=dict(radialaxis=dict(visible=True, range=[0,100])), showlegend=False)
            st.plotly_chart(plot_layout(fig, 330), use_container_width=True)
        else:
            st.info("Not enough assessed elements for a profile yet.")

    st.markdown('<div class="section-title">Students Requiring Attention</div>', unsafe_allow_html=True)
    attention = filtered[filtered["Performance Status"] != "On Track"].copy()
    cols = ["Register No", "Student Name", "Attendance %", "Communication Score %", "Aptitude Score %"]
    if programme == "B.Com": cols += ["Practical Stock Market Score %", "Financial Data Analytics Score %"]
    else: cols += ["DSA Skill Development", "DSA Extra Questions Skill", "DSA Skill Development %", "DSA Extra Questions Skill %"]
    cols += ["Composite Performance %", "Performance Status"]
    cols = [c for c in cols if c in attention.columns]
    if attention.empty:
        st.success("No students in the current filter require attention.")
    else:
        st.dataframe(attention[cols], use_container_width=True, hide_index=True)
    benchmark_explainer(programme)

def _student_name(master, reg):
    if not reg:
        return ""
    x = master[master["Register No"].astype(str) == str(reg)]
    return text(x.iloc[0].get("Student Name", "")) if not x.empty else ""


def attendance_tab(filtered, data, selected_reg):
    if "Attendance %" not in filtered.columns or filtered["Attendance %"].dropna().empty:
        st.info("Attendance data is not available for the current filter.")
        return
    master = data["master"]
    if not selected_reg:
        c1, c2, c3, c4 = st.columns(4)
        vals = filtered["Attendance %"].dropna()
        c1.metric("Average Attendance", f"{vals.mean():.1f}%")
        c2.metric("Highest", f"{vals.max():.1f}%")
        c3.metric(f"At Risk (<{ATTENDANCE_AT_RISK:.0f}%)", int((vals < ATTENDANCE_AT_RISK).sum()))
        c4.metric(f"Needs Attention", int(((vals >= ATTENDANCE_AT_RISK) & (vals < ATTENDANCE_ON_TRACK)).sum()))
        x = filtered[["Student Name", "Attendance %"]].dropna().sort_values("Attendance %")
        fig = px.bar(x, x="Attendance %", y="Student Name", orientation="h", color_discrete_sequence=[TEAL])
        fig.add_vline(x=ATTENDANCE_AT_RISK, line_dash="dash", line_color=PINK, annotation_text=f"{ATTENDANCE_AT_RISK:.0f}% At Risk")
        fig.add_vline(x=ATTENDANCE_ON_TRACK, line_dash="dash", line_color=TEAL_DARK, annotation_text=f"{ATTENDANCE_ON_TRACK:.0f}% On Track")
        st.plotly_chart(plot_layout(fig, max(360, len(x) * 28)), use_container_width=True)
        st.info("Select a student in the sidebar to view present days, absent days, total days, class comparison and the complete day-wise record.")
        return

    att = data["attendance"]
    row = att[att["Register No"].astype(str) == str(selected_reg)]
    if row.empty:
        st.info("No attendance record is available for this student.")
        return
    r = row.iloc[0]
    student_row = master[master["Register No"].astype(str) == str(selected_reg)].iloc[0]
    st.markdown(f'<div class="section-title">{r.get("Student Name","")} - Attendance Details</div>', unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total Days", str(int(r["Total Days"])) if pd.notna(r.get("Total Days")) else "N/A")
    c2.metric("Present Days", str(int(r["Present Days"])) if pd.notna(r.get("Present Days")) else "N/A")
    c3.metric("Absent Days", str(int(r["Absent Days"])) if pd.notna(r.get("Absent Days")) else "N/A")
    c4.metric("Attendance %", display_value(r.get("Attendance %", np.nan), "percent"))
    info = []
    if text(r.get("Roll No", "")): info.append(f"Roll No: **{r.get('Roll No')}**")
    if text(r.get("Department", "")): info.append(f"Department: **{r.get('Department')}**")
    if info: st.markdown(" &nbsp; • &nbsp; ".join(info), unsafe_allow_html=True)

    benchmark_comparison("Attendance", "Attendance %", student_row, master, TEAL, is_attendance=True)
    pctv = r.get("Attendance %", np.nan)
    if pd.notna(pctv) and pctv < ATTENDANCE_AT_RISK:
        st.error(f"Mentor insight: Attendance is {pctv:.1f}%, below the {ATTENDANCE_AT_RISK:.0f}% At Risk benchmark. Immediate follow-up is recommended.")
    elif pd.notna(pctv) and pctv < ATTENDANCE_ON_TRACK:
        st.warning(f"Mentor insight: Attendance is {pctv:.1f}%, within the monitoring band. Target at least {ATTENDANCE_ON_TRACK:.0f}%.")
    elif pd.notna(pctv):
        st.success(f"Mentor insight: Attendance is {pctv:.1f}% and meets the On Track benchmark.")

    daily = data.get("attendance_daily", pd.DataFrame())
    daily = daily[daily["Register No"].astype(str) == str(selected_reg)].copy() if not daily.empty else daily
    if not daily.empty:
        absent = daily[daily["Status"].astype(str).str.lower().isin(["ab", "abs", "absent", "a"])]
        st.markdown("#### Absence Summary")
        if absent.empty:
            st.success("No absence entries are recorded in the available day-wise data.")
        else:
            c1,c2 = st.columns([.25,.75])
            c1.metric("Recorded Absence Days", len(absent))
            c2.dataframe(absent[["Date / Session", "Status"]], use_container_width=True, hide_index=True)
        with st.expander("View complete day-wise attendance"):
            st.dataframe(daily[["Date / Session", "Status"]], use_container_width=True, hide_index=True)
    benchmark_explainer(data["programme"])

def communication_tab(filtered, data, selected_reg):
    if "Communication Score %" not in filtered.columns or filtered["Communication Score %"].dropna().empty:
        st.info("Communication data is not available for the current filter.")
        return
    master = data["master"]
    if not selected_reg:
        a, b = st.columns([1.3, .7])
        with a:
            x = filtered[["Student Name", "Communication Score %"]].dropna().sort_values("Communication Score %")
            fig = px.bar(x, x="Communication Score %", y="Student Name", orientation="h", color_discrete_sequence=[YELLOW])
            fig.add_vline(x=SKILL_CRITICAL, line_dash="dash", line_color=PINK, annotation_text="40% Critical")
            fig.add_vline(x=SKILL_ON_TRACK, line_dash="dash", line_color=TEAL_DARK, annotation_text="60% On Track")
            st.plotly_chart(plot_layout(fig, max(350, len(x) * 26)), use_container_width=True)
        with b:
            if "Latest CEFR" in filtered.columns:
                c = filtered["Latest CEFR"].replace("", np.nan).dropna().value_counts().reset_index(); c.columns=["CEFR","Students"]
                if not c.empty:
                    fig=px.pie(c,names="CEFR",values="Students",hole=.45,color_discrete_sequence=[TEAL,YELLOW,PINK,BLUE,PURPLE]); st.plotly_chart(plot_layout(fig,350),use_container_width=True)
        st.info("Select a student in the sidebar to view class comparison, activity-wise rubric scores, totals and CEFR progression.")
        return

    student_row = master[master["Register No"].astype(str) == str(selected_reg)].iloc[0]
    detail = data.get("communication_detail", pd.DataFrame())
    detail = detail[detail["Register No"].astype(str) == str(selected_reg)].copy() if not detail.empty else detail
    c1,c2,c3,c4=st.columns(4)
    c1.metric("Communication Score", display_value(student_row.get("Communication Score %", np.nan),"percent"))
    c2.metric("Latest CEFR", text(student_row.get("Latest CEFR","")) or "N/A")
    c3.metric("Activities Recorded", int(len(detail)) if not detail.empty else 0)
    c4.metric("Average Activity Total", f"{detail['Total'].mean():.1f}" if not detail.empty and detail["Total"].notna().any() else "N/A")
    benchmark_comparison("Communication", "Communication Score %", student_row, master, YELLOW)
    v=student_row.get("Communication Score %",np.nan)
    if pd.notna(v) and v<SKILL_CRITICAL: st.error(f"Mentor insight: Communication is {v:.1f}%, below the critical benchmark. Review rubric-level gaps and prioritise targeted practice.")
    elif pd.notna(v) and v<SKILL_ON_TRACK: st.warning(f"Mentor insight: Communication is {v:.1f}%. It needs attention until the student reaches the {SKILL_ON_TRACK:.0f}% On Track benchmark.")
    elif pd.notna(v): st.success(f"Mentor insight: Communication meets the {SKILL_ON_TRACK:.0f}% On Track benchmark. Review the rubric breakdown for further development.")
    if not detail.empty:
        st.markdown("#### Activity-wise Communication Performance")
        st.dataframe(detail[["Activity","Date / Context","Criteria Breakdown","Total","CEFR"]],use_container_width=True,hide_index=True)
    benchmark_explainer(data["programme"])

def aptitude_tab(filtered, data, selected_reg):
    if "Aptitude Score %" not in filtered.columns or filtered["Aptitude Score %"].dropna().empty:
        st.info("Aptitude data is not available for the current filter.")
        return
    master = data["master"]
    if not selected_reg:
        c1,c2,c3=st.columns(3)
        c1.metric("Average Aptitude",f"{filtered['Aptitude Score %'].mean():.1f}%")
        if "Aptitude Questions Solved" in filtered.columns: c2.metric("Avg Questions Solved",f"{filtered['Aptitude Questions Solved'].mean():.1f}")
        if "Aptitude Correct" in filtered.columns: c3.metric("Avg Correct Answers",f"{filtered['Aptitude Correct'].mean():.1f}")
        x=filtered[["Student Name","Aptitude Score %"]].dropna().sort_values("Aptitude Score %")
        fig=px.bar(x,x="Aptitude Score %",y="Student Name",orientation="h",color_discrete_sequence=[PINK]); fig.add_vline(x=SKILL_CRITICAL,line_dash="dash",line_color=PINK,annotation_text="40% Critical"); fig.add_vline(x=SKILL_ON_TRACK,line_dash="dash",line_color=TEAL_DARK,annotation_text="60% On Track")
        st.plotly_chart(plot_layout(fig,max(350,len(x)*26)),use_container_width=True)
        st.info("Select a student in the sidebar to view class comparison, solved/correct/wrong questions, accuracy, assessments and task-wise scores.")
        return

    r=master[master["Register No"].astype(str)==str(selected_reg)].iloc[0]
    summary=data.get("aptitude_summary_detail",pd.DataFrame()); sr=summary[summary["Register No"].astype(str)==str(selected_reg)] if not summary.empty else pd.DataFrame(); srow=sr.iloc[0] if not sr.empty else None
    c1,c2,c3,c4,c5=st.columns(5)
    c1.metric("Aptitude Score",display_value(r.get("Aptitude Score %",np.nan),"percent")); c2.metric("Questions Solved",display_value(srow.get("Questions Solved",np.nan) if srow is not None else r.get("Aptitude Questions Solved",np.nan),"count")); c3.metric("Correct",display_value(srow.get("Correct Answers",np.nan) if srow is not None else r.get("Aptitude Correct",np.nan),"count")); c4.metric("Wrong",display_value(srow.get("Wrong Answers",np.nan) if srow is not None else r.get("Aptitude Wrong",np.nan),"count")); acc=srow.get("Overall Accuracy %",np.nan) if srow is not None else r.get("Aptitude Accuracy %",np.nan); c5.metric("Accuracy",display_value(acc,"percent"))
    benchmark_comparison("Aptitude", "Aptitude Score %", r, master, PINK)
    v=r.get("Aptitude Score %",np.nan)
    if pd.notna(v) and v<SKILL_CRITICAL: st.error(f"Mentor insight: Aptitude is {v:.1f}%, below the critical benchmark. Review incorrect-answer patterns and weaker topics.")
    elif pd.notna(v) and v<SKILL_ON_TRACK: st.warning(f"Mentor insight: Aptitude is {v:.1f}% and needs attention. Use task-level results to focus practice.")
    elif pd.notna(v): st.success(f"Mentor insight: Aptitude meets the {SKILL_ON_TRACK:.0f}% On Track benchmark. Use task results to strengthen weaker topics.")
    ca=data.get("aptitude_ca_detail",pd.DataFrame()); ca=ca[ca["Register No"].astype(str)==str(selected_reg)].copy() if not ca.empty else ca
    tasks=data.get("aptitude_task_detail",pd.DataFrame()); tasks=tasks[tasks["Register No"].astype(str)==str(selected_reg)].copy() if not tasks.empty else tasks
    if not ca.empty:
        st.markdown("#### Continuous Assessment Details"); st.dataframe(ca[["Assessment","Date","Total Questions","Questions Solved","Correct Answers","Marks %","Assessment Status"]],use_container_width=True,hide_index=True)
    if not tasks.empty:
        st.markdown("#### Aptitude Task / Practice Details"); st.dataframe(tasks[["Date","Task / Assessment","Topic","Category","Score","Max Score","Status"]],use_container_width=True,hide_index=True)
    benchmark_explainer(data["programme"])

def bcom_skill_tab(filtered, data, selected_reg, label, field, color, detail_key, report_key):
    if field not in filtered.columns or filtered[field].dropna().empty:
        st.info(f"{label} data is not available for the current filter.")
        return
    master=data["master"]
    if not selected_reg:
        vals=filtered[field].dropna(); c1,c2,c3=st.columns(3); c1.metric(f"Average {label}",f"{vals.mean():.1f}%"); c2.metric("Highest",f"{vals.max():.1f}%"); c3.metric("Students Assessed",int(vals.count()))
        x=filtered[["Student Name",field]].dropna().sort_values(field); fig=px.bar(x,x=field,y="Student Name",orientation="h",color_discrete_sequence=[color]); fig.add_vline(x=SKILL_CRITICAL,line_dash="dash",line_color=PINK,annotation_text="40% Critical"); fig.add_vline(x=SKILL_ON_TRACK,line_dash="dash",line_color=TEAL_DARK,annotation_text="60% On Track"); st.plotly_chart(plot_layout(fig,max(350,len(x)*26)),use_container_width=True)
        st.info("Select a student in the sidebar to view class comparison, session-wise activities, marks and submitted report links/files.")
        return
    r=master[master["Register No"].astype(str)==str(selected_reg)].iloc[0]
    detail=data.get(detail_key,pd.DataFrame()); detail=detail[detail["Register No"].astype(str)==str(selected_reg)].copy() if not detail.empty else detail
    reports=data.get(report_key,pd.DataFrame()); reports=reports[reports["Register No"].astype(str)==str(selected_reg)].copy() if not reports.empty else reports
    vals=detail["Score / 5"].dropna() if not detail.empty else pd.Series(dtype=float)
    c1,c2,c3,c4=st.columns(4); c1.metric(f"{label} Score",display_value(r.get(field,np.nan),"percent")); c2.metric("Sessions Assessed",int(vals.count())); c3.metric("Average Mark / 5",f"{vals.mean():.1f}" if not vals.empty else "N/A"); c4.metric("Highest Mark / 5",f"{vals.max():.1f}" if not vals.empty else "N/A")
    benchmark_comparison(label, field, r, master, color)
    v=r.get(field,np.nan)
    if pd.notna(v) and v<SKILL_CRITICAL: st.error(f"Mentor insight: {label} is {v:.1f}%, below the critical benchmark. Review low-scoring sessions and missing submissions.")
    elif pd.notna(v) and v<SKILL_ON_TRACK: st.warning(f"Mentor insight: {label} is {v:.1f}% and needs attention. Review session-level scores and incomplete submissions.")
    elif pd.notna(v): st.success(f"Mentor insight: {label} meets the On Track benchmark. Continue reviewing session-level consistency.")
    if not detail.empty:
        st.markdown("#### Session-wise Performance"); st.dataframe(detail[["Group","Week","Session","Activity","Score / 5"]],use_container_width=True,hide_index=True)
    if not reports.empty:
        with st.expander("Submitted Reports / Links"): st.dataframe(reports[["Context","Report Link / File"]],use_container_width=True,hide_index=True)
    benchmark_explainer(data["programme"])

def dsa_skill_tab(filtered, data, selected_reg, label, field, color, summary_key, detail_key):
    if field not in filtered.columns or filtered[field].dropna().empty:
        st.info(f"{label} data is not available for the current filter.")
        return
    master=data["master"]
    progress_field = "DSA Skill Development %" if field == "DSA Skill Development" else "DSA Extra Questions Skill %"
    assigned_field = f"{field} Assigned"
    if not selected_reg:
        vals=filtered[field].dropna(); progress=filtered[progress_field].dropna() if progress_field in filtered.columns else pd.Series(dtype=float)
        c1,c2,c3,c4=st.columns(4); c1.metric(f"Average {label}",f"{vals.mean():.1f}"); c2.metric("Average Progress",f"{progress.mean():.1f}%" if not progress.empty else "N/A"); c3.metric("Highest Completed",f"{vals.max():.0f}"); c4.metric("Students Assessed",int(vals.count()))
        x=filtered[["Student Name",progress_field]].dropna().sort_values(progress_field) if progress_field in filtered.columns else pd.DataFrame()
        if not x.empty:
            fig=px.bar(x,x=progress_field,y="Student Name",orientation="h",color_discrete_sequence=[color]); fig.add_vline(x=SKILL_CRITICAL,line_dash="dash",line_color=PINK,annotation_text="40% Critical"); fig.add_vline(x=SKILL_ON_TRACK,line_dash="dash",line_color=TEAL_DARK,annotation_text="60% On Track"); st.plotly_chart(plot_layout(fig,max(350,len(x)*26)),use_container_width=True)
        st.caption("Progress % = Questions Completed / Total Questions Assigned × 100. It does not depend on the top student's score.")
        st.info("Select a student in the sidebar to view assigned/completed questions, progress comparison, question status, class-work marks and profile details.")
        return
    r=master[master["Register No"].astype(str)==str(selected_reg)].iloc[0]
    summary=data.get(summary_key,pd.DataFrame()); sr=summary[summary["Register No"].astype(str)==str(selected_reg)] if not summary.empty else pd.DataFrame()
    detail=data.get(detail_key,pd.DataFrame()); detail=detail[detail["Register No"].astype(str)==str(selected_reg)].copy() if not detail.empty else detail
    completed=int(detail["Status"].astype(str).str.lower().eq("completed").sum()) if not detail.empty else 0; not_completed=int(detail["Status"].astype(str).str.lower().eq("not completed").sum()) if not detail.empty else 0; not_assessed=int(detail["Status"].astype(str).str.lower().eq("not assessed").sum()) if not detail.empty else 0
    assigned=r.get(assigned_field,np.nan); progress=r.get(progress_field,np.nan); avg_mark=detail["Class Work / 5"].mean() if not detail.empty and detail["Class Work / 5"].notna().any() else np.nan
    c1,c2,c3,c4,c5=st.columns(5); c1.metric("Questions Assigned",str(int(assigned)) if pd.notna(assigned) else "N/A"); c2.metric("Questions Completed",str(int(r.get(field))) if pd.notna(r.get(field,np.nan)) else "N/A"); c3.metric("Progress",display_value(progress,"percent")); c4.metric("Not Completed",not_completed); c5.metric("Avg Class Work / 5",f"{avg_mark:.1f}" if pd.notna(avg_mark) else "N/A")
    benchmark_comparison(label + " Progress", progress_field, r, master, color)
    if pd.notna(progress) and progress<SKILL_CRITICAL: st.error(f"Mentor insight: {label} progress is {progress:.1f}%, below the critical benchmark. Prioritise pending questions.")
    elif pd.notna(progress) and progress<SKILL_ON_TRACK: st.warning(f"Mentor insight: {label} progress is {progress:.1f}% and needs attention. Review pending questions and class-work scores.")
    elif pd.notna(progress): st.success(f"Mentor insight: {label} progress meets the On Track benchmark. Continue monitoring pending items and class-work quality.")
    if not sr.empty:
        profile=text(sr.iloc[0].get("HackerRank Profile",""));
        if profile: st.caption(f"HackerRank profile: {profile}")
    if not detail.empty:
        st.markdown("#### Question-wise DSA Performance"); st.dataframe(detail[["Week","Question / Topic","Date / Context","Status","Class Work / 5","Question Link"]],use_container_width=True,hide_index=True)
    benchmark_explainer(data["programme"])

def student_view(filtered, master, programme, data=None, source_mode="Local Excel", selected_reg=None):
    if selected_reg:
        reg = str(selected_reg)
    else:
        pool = filtered if not filtered.empty else master
        options = pool[["Register No", "Student Name"]].drop_duplicates().sort_values("Student Name")
        if options.empty:
            st.info("No student is available for the current filter.")
            return
        labels=[f"{r['Student Name']} • {r['Register No']}" for _,r in options.iterrows()]
        chosen=st.selectbox("Select student for detailed view",labels,key="student_detail"); reg=chosen.rsplit("•",1)[-1].strip()
    x=master[master["Register No"].astype(str)==reg]
    if x.empty: st.info("Student record is not available."); return
    r=x.iloc[0]; status=r.get("Performance Status",""); sc=status_color(status)
    st.markdown(f'''<div class="student-banner"><div style="display:flex;justify-content:space-between;gap:12px;align-items:flex-start;flex-wrap:wrap;"><div><div style="font-size:1.5rem;font-weight:800;color:{TEXT};">{r.get('Student Name','')}</div><div style="color:{MUTED};margin-top:3px;">Register No: {reg} · Programme: {programme}</div></div><span class="status-pill" style="background:{sc}22;color:{sc};font-size:.88rem;">{status.upper()}</span></div></div>''',unsafe_allow_html=True)
    cols=st.columns(5)
    for col,(label,field,kind,color) in zip(cols,metric_specs(programme)):
        with col: st.markdown(card_html(label,r.get(field,np.nan),kind,color,"Student Performance"),unsafe_allow_html=True)
    render_alert_panel(r,programme)
    st.markdown("#### Status Rationale")
    for reason in performance_reasons(r,programme): st.markdown(f"- {reason}")
    comp=comparison_data(r,master,programme); st.markdown("#### Student vs Class Benchmark"); st.dataframe(comp,use_container_width=True,hide_index=True)
    specs=percent_specs(programme); cats=[]; vals=[]
    for label,field in specs:
        v=r.get(field,np.nan)
        if pd.notna(v): cats.append(label); vals.append(v)
    if len(vals)>=3:
        fig=go.Figure(go.Scatterpolar(r=vals+[vals[0]],theta=cats+[cats[0]],fill="toself",line=dict(color=TEAL_DARK),fillcolor="rgba(86,189,177,.20)")); fig.update_layout(polar=dict(radialaxis=dict(visible=True,range=[0,100])),showlegend=False); st.plotly_chart(plot_layout(fig,390),use_container_width=True)
    benchmark_explainer(programme)
    if data is not None:
        try:
            pdf=build_student_pdf(r,master,data,programme,source_mode); st.download_button("Download Student Performance Report (PDF)",pdf,f"{text(r.get('Student Name','Student')).replace(' ','_')}_performance_report.pdf","application/pdf", key=f"student_pdf_{reg}")
        except ModuleNotFoundError: st.warning("PDF report support requires ReportLab. Run: pip install reportlab")

def data_table(filtered, programme):
    cols = ["Register No", "Student Name", "Performance Status", "Composite Performance %"]
    cols += [field for _, field, _, _ in metric_specs(programme)]
    cols = list(dict.fromkeys(c for c in cols if c in filtered.columns))
    view = filtered[cols].copy()
    st.dataframe(view, use_container_width=True, hide_index=True)
    st.download_button(
        "Download filtered report (CSV)",
        view.to_csv(index=False).encode("utf-8"),
        f"{programme.replace('.', '').replace(' ', '_')}_skill_tracker_report.csv",
        "text/csv",
        use_container_width=False,
        key=f"csv_export_{programme}",
    )


# -------------------- MAIN APP --------------------
def find_local_file(programme):
    data_dir = Path(__file__).parent / "data"
    if not data_dir.exists():
        return None
    files = sorted(data_dir.glob("*.xlsx"))
    keywords = LOCAL_FILE_KEYWORDS[programme]
    for f in files:
        if any(k in f.name.lower() for k in keywords):
            return f
    return None


def main():
    with st.sidebar:
        st.markdown(f"## {APP_TITLE}")
        programme_label = st.selectbox("Programme", list(PROGRAMMES.keys()))
        programme = PROGRAMMES[programme_label]
        source_options = ["Local Excel", "Upload Excel"]
        if google_secrets_available(): source_options.insert(0, "Live Google Sheets")
        source_mode = st.radio("Data Source", source_options)

    data = None
    connection_ok = True
    if source_mode == "Live Google Sheets":
        try:
            ids = st.secrets["google_sheets"]
            spreadsheet_id = ids["bcom_spreadsheet_id"] if programme == "B.Com" else ids["bca_spreadsheet_id"]
            data = read_google_tracker(spreadsheet_id, programme)
            st.session_state[f"last_google_data_{programme}"] = data
        except ModuleNotFoundError:
            st.error("Google Sheets packages are missing. Run: pip install gspread google-auth")
            st.stop()
        except Exception as e:
            connection_ok = False
            backup = st.session_state.get(f"last_google_data_{programme}")
            if backup is not None:
                data = backup
                st.warning(f"Live Google Sheets could not be refreshed. Showing the last data loaded in this browser session. Error: {e}")
            else:
                st.error(f"Could not load Google Sheets: {e}")
                st.info("Check the spreadsheet ID, service-account credentials, APIs, and whether the sheet is shared with the service-account email.")
                st.stop()
    elif source_mode == "Local Excel":
        f=find_local_file(programme)
        if f is None: st.error(f"Could not find the {programme} Excel tracker inside the data folder."); st.stop()
        data=read_local_tracker(str(f),os.path.getmtime(f),programme)
    else:
        with st.sidebar: up=st.file_uploader(f"Upload {programme} tracker",type=["xlsx"])
        if up is None: st.info("Upload the tracker from the sidebar to continue."); st.stop()
        data=read_uploaded_tracker(up.getvalue(),up.name,programme)

    master=data["master"]
    if master.empty: st.error("Student data could not be identified. Check the tracker sheet names and structure."); st.stop()

    filtered, selected_reg, active_filters = apply_filters(master)
    if filtered.empty:
        hero(programme_label,data,filtered,source_mode,connection_ok); st.warning("No students match the selected filters. Change or reset the filters."); st.stop()

    with st.sidebar:
        st.divider(); st.markdown("### Data Refresh")
        if source_mode=="Live Google Sheets":
            if connection_ok: st.success("LIVE · Google Sheets connected")
            else: st.error("Data connection issue · last available data")
            st.caption(f"Automatic refresh every {REFRESH_SECONDS} seconds.")
        else: st.caption(f"The tracker is checked every {REFRESH_SECONDS} seconds.")
        if st.button("Refresh data now",use_container_width=True): st.cache_data.clear(); st.rerun()

    hero(programme_label,data,filtered,source_mode,connection_ok)
    top_cards(filtered,master,programme,selected_reg,active_filters)
    st.write("")

    if programme=="B.Com": names=["Overview","Attendance","Communication","Aptitude","Practical Stock","Financial Data Analytics","Student View","Data Table"]
    else: names=["Overview","Attendance","Communication","Aptitude","DSA Skill Development","DSA Extra Questions","Student View","Data Table"]
    tabs=st.tabs(names)
    with tabs[0]: overview_tab(filtered,programme,data,selected_reg,source_mode)
    with tabs[1]: attendance_tab(filtered,data,selected_reg)
    with tabs[2]: communication_tab(filtered,data,selected_reg)
    with tabs[3]: aptitude_tab(filtered,data,selected_reg)
    with tabs[4]:
        if programme=="B.Com": bcom_skill_tab(filtered,data,selected_reg,"Practical Stock Market","Practical Stock Market Score %",BLUE,"practical_detail","practical_reports")
        else: dsa_skill_tab(filtered,data,selected_reg,"DSA Skill Development","DSA Skill Development",BLUE,"dsa_core_summary","dsa_core_detail")
    with tabs[5]:
        if programme=="B.Com": bcom_skill_tab(filtered,data,selected_reg,"Financial Data Analytics","Financial Data Analytics Score %",PURPLE,"financial_detail","financial_reports")
        else: dsa_skill_tab(filtered,data,selected_reg,"DSA Extra Questions Skill","DSA Extra Questions Skill",PURPLE,"dsa_extra_summary","dsa_extra_detail")
    with tabs[6]: student_view(filtered,master,programme,data,source_mode,selected_reg)
    with tabs[7]: data_table(filtered,programme)

    st.markdown("---")
    st.markdown('<span class="small-note">Composite Performance is a dashboard analytical indicator. Missing/unassessed values are shown as N/A and are not treated as zero. Monitoring thresholds are configurable in the app and are not official benchmarks stored in Google Sheets.</span>',unsafe_allow_html=True)



if __name__ == "__main__":
    main()
